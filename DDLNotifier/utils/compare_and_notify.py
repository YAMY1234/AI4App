import os
from datetime import datetime
import pandas as pd
from DDLNotifier.email_sender import send_email
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from DDLNotifier.config import CONFIG
from filelock import FileLock

recipient_email = CONFIG.RECIPEINT_EMAIL


def update_error_urls_with_old_data(new_data, old_data):
    """
    Update URLs in new_data that start with 'Error processing the page:'
    with corresponding URLs from old_data if available.

    :param new_data: DataFrame with new program data
    :param old_data: DataFrame with old program data
    :return: Updated new_data DataFrame
    """
    for index, row in new_data.iterrows():
        program_name = row['Programme']
        url_link = row['Deadline']

        if isinstance(url_link, str) and url_link.startswith("Error processing the page:"):
            # Find the matching ProgramName in old_data
            old_row = old_data[old_data['Programme'] == program_name]

            if not old_row.empty:
                # Replace with the URL Link from old_data if it exists
                old_url = old_row.iloc[0]['Deadline']
                new_data.at[index, 'Deadline'] = old_url
            else:
                # Keep the original 'Error processing the page:' content
                print(f"No matching ProgramName in old data for {program_name}. Keeping original error message.")

    return new_data

def compare_and_notify(old_data, new_data, log_path, school_name):
    """
    Compare old and new data, and log differences.

    Parameters:
    - old_data: DataFrame containing the previous data (identical to the outer scope variable).
    - new_data: DataFrame containing the current data to compare against old_data.
    - log_path: Path to the log file where results are recorded.
    - school_name: Name of the school for which the data is compared.

    This function assumes that 'old_data' from the outer scope is passed to it directly.
    """
    with open(log_path, "a") as log_file:
        log_file.write(f"Function called at {datetime.now()}\n")

    if old_data.empty:
        print("No old data to compare with. Saving new data.")
        return

    # Check for any differences
    if not old_data.equals(new_data):
        print("Data differences detected...")

        changes_detected = []
        new_rows_detected = []
        deleted_rows_detected = []

        for index, new_row in new_data.iterrows():
            # Check if the row exists in the old data
            old_row = old_data.loc[old_data['Programme'] == new_row['Programme']]

            # If the row exists, check for deadline changes
            if not old_row.empty:
                if old_row['Deadline'].values[0] != new_row['Deadline']:
                    changes_detected.append({
                        'Programme': new_row['Programme'],
                        'Old Deadline': old_row['Deadline'].values[0],
                        'New Deadline': new_row['Deadline']
                    })
            else:
                # If the row does not exist in old data, it's a new addition
                new_rows_detected.append(new_row)

        # Check for deleted programmes
        for index, old_row in old_data.iterrows():
            new_row = new_data.loc[new_data['Programme'] == old_row['Programme']]
            if new_row.empty:
                deleted_rows_detected.append(old_row)

        # Save changes to an email if there are any changes
        save_changes_to_email(changes_detected, new_rows_detected, deleted_rows_detected,
                              school_name, log_path)

        # Save changes to an Excel file if there are any changes
        save_changes_to_excel(changes_detected, new_rows_detected, deleted_rows_detected,
                              school_name)
    else:
        print("No changes detected in the data content.")


def save_changes_to_email(changes_detected, new_rows_detected, deleted_rows_detected, m_school_name, log_path):
    # Preparing email content
    subject = f"Changes Detected in School: {m_school_name}'s Programmes"
    body = ""

    if changes_detected:
        body += "Deadline changes detected:\n\n"
        for change in changes_detected:
            body += (
                f"School: {m_school_name}, Programme: {change['Programme']}\n"
                f"Old Deadline: {change['Old Deadline']}\n"
                f"New Deadline: {change['New Deadline']}\n\n")

    if new_rows_detected:
        body += "New programmes added:\n\n"
        for new_row in new_rows_detected:
            body += (
                f"School: {m_school_name}, Programme: {new_row['Programme']}\n"
                f"Deadline: {new_row['Deadline']}\n\n")

    if deleted_rows_detected:
        body += "Programmes deleted:\n\n"
        for del_row in deleted_rows_detected:
            body += f"School: {m_school_name}, Programme: {del_row['Programme']}\n\n"

    # Sending the email if there are any changes
    if changes_detected or new_rows_detected or deleted_rows_detected:
        send_email(subject, body, recipient_email)
        with open(log_path, "a") as log_file:
            log_file.write(f"Email sent: {subject} | {body}\n")
        print("Email notification sent for the detected changes.")
    else:
        print("No changes detected.")


def save_changes_to_excel(changes_detected, new_rows_detected, deleted_rows_detected, school_name):
    # Current timestamp (year, month, day, hour)
    now = datetime.now()
    timestamp = now.strftime("%Y-%m-%d %H")
    file_name = f"{timestamp}-changes.xlsx"
    file_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), file_name)

    # Define headers and styles
    headers = ["Alert Label", "School Name", "Program Name", "Old Deadline", "New Deadline",
               "Record Time"]
    header_font = Font(bold=True)
    styles = {
        "Added Program": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Deleted Program": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "Deadline Change": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    }

    # Using file lock to ensure atomicity in file operations
    lock = FileLock(f"{file_path}.lock")
    with lock:
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
        else:
            wb = Workbook()

        # Check if the sheet for this school already exists, if not, create a new one
        if school_name in wb.sheetnames:
            ws = wb[school_name]
        else:
            ws = wb.create_sheet(school_name)
            ws.append(headers)
            for cell in ws["1:1"]:
                cell.font = header_font

        # Add data to worksheet
        for change in changes_detected:
            ws.append(["Deadline Change", school_name, change['Programme'], change['Old Deadline'],
                       change['New Deadline'], timestamp])

        for new_row in new_rows_detected:
            ws.append(["Added Program", school_name, new_row['Programme'], "", new_row['Deadline'],
                       timestamp])

        for del_row in deleted_rows_detected:
            ws.append(
                ["Deleted Program", school_name, del_row['Programme'], del_row['Deadline'], "",
                 timestamp])

        # Apply styles to rows based on Alert Label
        for row in ws.iter_rows(min_row=2, max_col=6):
            if row[0].value in styles:
                for cell in row:
                    cell.fill = styles[row[0].value]

        # Save the workbook
        wb.save(file_path)
        print(f"Excel file saved to {file_path}")

# Example usage
if __name__ == "__main__":
    import pandas as pd

    # Mock old data
    old_data = pd.DataFrame({
        'Programme': ['Program A', 'Program B', 'Program C'],
        'Deadline': ['2023-12-01', '2023-11-15', '2023-12-20']
    })

    # Mock new data
    new_data = pd.DataFrame({
        'Programme': ['Program A', 'Program B', 'Program D'],
        'Deadline': ['2023-12-05', '2023-11-15', '2023-12-25']
    })

    # Define log path and school name
    log_path = 'test_log.txt'
    school_name = 'Test School'

    # Run the compare_and_notify function with mock data
    compare_and_notify(old_data, new_data, log_path, school_name)

    print("Test completed. Check the local directory for the Excel file.")
