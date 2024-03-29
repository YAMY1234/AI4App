import json
import os
import re
import threading
import time

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
import concurrent.futures
from config.program_details_header import header

from tools.general import request_with_retry
from tools.general import translate_text


class BaseProgramURLCrawler:
    def __init__(self, base_url, Uni_ID, Uni_name=None, verbose=False):
        self.base_url = base_url
        self.data_folder = "data"
        self.Uni_ID = Uni_ID
        if not Uni_name:
            self.Uni_name = Uni_ID
        else:
            self.Uni_name = Uni_name
        self.verbose = verbose

        if not os.path.exists(self.data_folder):
            os.makedirs(self.data_folder)


    def crawl(self):
        if self.verbose:
            print(f"crawling program urls for {self.Uni_name}...")
        program_words = self._load_program_words()
        soup = self._fetch_html(self.base_url)
        program_url_pairs = self._parse_programs(soup, program_words)
        self._store_results(program_url_pairs)

    def _load_program_words(self):
        # This can be modified if the source file name or structure changes
        wb = load_workbook(os.path.join(
            self.data_folder, 'program_word_bank.xlsx'))
        sheet = wb.active
        return [row[0].lower() for row in sheet.iter_rows(min_row=2, values_only=True)]

    def _fetch_html(self, url):
        response = requests.get(url)
        response.raise_for_status()
        return BeautifulSoup(response.text, 'html.parser')

    def _parse_programs(self, soup, program_words):
        raise NotImplementedError(
            "Each specific crawler should implement this method!")

    def _store_results(self, program_url_pairs):
        Uni_ID = self.Uni_ID
        school_data_folder = os.path.join(self.data_folder, Uni_ID)

        if not os.path.exists(school_data_folder):
            os.makedirs(school_data_folder)

        # Create a new workbook and worksheet
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = "Program URLs"

        # Add header row
        new_ws.append(["Program Name", "URL", "Faculty"])

        # Store the results in the excel sheet
        for program_name, details in program_url_pairs.items():
            url, faculty = details
            new_ws.append([program_name, url, faculty])

        # Save the file
        try:
            new_wb.save(os.path.join(
                school_data_folder, 'program_url_pair.xlsx'))
        except PermissionError:
            # save the file with a time stamp: YYYY-MM-DD_HH-MM-SS
            new_wb.save(os.path.join(
                school_data_folder, f'program_url_pair_{time.strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'))
            print(
                f"PermissionError: failed to save program_url_pair.xlsx. Saved as program_url_pair_{time.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx instead.")

        return


intervew_keywords = ['interview', 'special qualifying examination', 'qualifying essay',
            'qualifying assessment', 'written examination', 'oral examination',
            'oral test', 'required to pass a test', 'written examination', 'written test', 'written assessment',
            'pieces of writing']

portfolio_keywords = ['written work', 'portfolio', "writing sample"]

maybe_keywords = ["may", "preferably", 'minimum of']

required_keywords = ['minimum of', 'at least', 'Ideally', 'is essential', 'must have', 'normally have',
                    'should also have', 'should have', 'need to have']

word_experience_keywords = ['work experience', 'professional experience', 'industrial experience',
                   'existing engineering and design skills', 'practical experience',
                   'experience as a professional', ' who can demonstrate aptitude and experience',
                   'years of relevant experience', 'clinical experience', 'teaching experience',
                   'years\' experience working', 'years of training in',
                   'extensive experience', 'relevant experience',
                   'relevant quantitative or qualitative research experience',
                   'experience working', 'professional involvement',
                   'with equivalent experience', 'industry experience', 'relevant work',
                   'field experience', 'relevant employment']
class BaseProgramDetailsCrawler:
    def __init__(self, Uni_ID, Uni_name=None, test=False, verbose=True):
        self.Uni_ID = Uni_ID
        if not Uni_name:
            self.Uni_name = Uni_ID
        else:
            self.Uni_name = Uni_name
        self.test = test
        self.verbose = verbose

        self.interview_keywords = intervew_keywords
        self.portfolio_keywords = portfolio_keywords
        self.maybe_keywords = maybe_keywords
        self.required_keywords = required_keywords
        self.word_experience_keywords = word_experience_keywords

        if self.test:
            self.useful_link_path = f'data/{self.Uni_ID}/program_useful_link_sample.xlsx'
            self.details_path = f'data/{self.Uni_ID}/program_details_sample.xlsx'
        else:
            self.useful_link_path = f'data/{self.Uni_ID}/program_useful_link.xlsx'
            self.details_path = f'data/{self.Uni_ID}/program_details.xlsx'

    def read_excel(self, path):
        return load_workbook(path)

    def get_program_useful_links(self):
        url_pair_path = f'data/{self.Uni_ID}/program_url_pair.xlsx'
        if self.test:
            url_pair_path = f'data/{self.Uni_ID}/program_url_pair_sample.xlsx'
            # generate a sample based on first 10 rows of program_url_pair.xlsx
            if not os.path.exists(url_pair_path):
                original_ = f'data/{self.Uni_ID}/program_url_pair.xlsx'
                wb = self.read_excel(original_)
                sheet = wb.active
                new_wb = Workbook()
                new_sheet = new_wb.active
                new_sheet.append(["Program Name", "URL", "Faculty"])
                for row in sheet.iter_rows(min_row=2, max_row=11, values_only=True):
                    new_sheet.append(row)
                new_wb.save(url_pair_path)

        if self.verbose:
            print(f"generating useful links for {self.Uni_name}...")
            print("reading url pairs from " + url_pair_path)
        wb = self.read_excel(url_pair_path)
        sheet = wb.active

        new_wb = Workbook()
        new_sheet = new_wb.active
        new_sheet.append(['项目名', '项目链接', header.college, '项目简介', '链接'])

        # Todo: adjust max_workers based on the number of cores
        with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
            results = list(executor.map(self.process_row, sheet.iter_rows(
                min_row=2, values_only=True), chunksize=1))

        for result in results:
            new_sheet.append(result)
        try:
            new_wb.save(self.useful_link_path)
            print("useful links saved to " + self.useful_link_path)
        except PermissionError:
            # save the file with a time stamp: YYYY-MM-DD_HH-MM-SS
            new_wb.save(os.path.join(
                f'data/{self.Uni_ID}/program_useful_link_{time.strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'))
            print(
                f"PermissionError: failed to save program_useful_link.xlsx. Saved as program_useful_link_{time.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx instead.")

    def process_row(self, row):
        link_bank = self.read_excel('data/url_bank.xlsx').active
        program_name, url, faculty = row

        response = requests.get(url)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, 'html.parser')
        intro_div = soup.find('div', class_='zero')
        intro_text = intro_div.get_text() if intro_div else ""

        useful_links = []
        for link in soup.find_all('a'):
            link_name = link.get_text().strip()
            link_url = link.get('href')

            for bank_row in link_bank.iter_rows(min_row=2, values_only=True):
                if bank_row[0].lower() in link_name.lower():
                    useful_links.append(f'{link_name}:{link_url}')
                    break

        return [program_name, url, faculty, intro_text] + useful_links

    def thread_task(self, row, headers, new_sheet, lock):

        program_name, program_url, program_faculty, intro_text, *useful_links = row
        program_details = {header.major: program_name,
                           header.college: program_faculty,
                           header.website_link: program_url,
                           header.project_intro: intro_text,
                           }

        self.get_data_from_main_url(program_url, program_details)
        # self.get_data_from_url(useful_links, program_details)

        # 根据 headers 生成要追加的行
        data_row = [program_details.get(header, '') for header in headers]
        if self.verbose:
            print(json.dumps(program_details, indent=4, ensure_ascii=False))
        else:
            print(f"succeeded with program: {program_name}.")

        with lock:  # 确保线程安全
            new_sheet.append(data_row)

    def generate_program_details(self, verbose=True, translate=False):
        if verbose:
            print(f"crawling program details for {self.Uni_name}...")
        self.scrape_program_details()
        self.write_program_constants()
        if translate:
            self.translate()
        if verbose:
            print(f"{self.Uni_ID} updated successfully.")

    def write_program_constants(self):
        # 从university_constants.csv中读取headers和数据
        with open('data/Constants/university_constants.csv', 'r', encoding='utf-8') as file:
            lines = [line.strip() for line in file.readlines()]
            headers = lines[0].split(sep=',')
            # All the following lines are data, each line is a university
            data = [line.split(sep=',') for line in lines[1:]]

        # open program_details.xlsx
        wb = self.read_excel(self.details_path)
        sheet = wb.active

        # 获取现有的headers
        existing_headers = [cell.value for cell in sheet[1]]

        # get data for current university
        data = [row for row in data if row[0] == self.Uni_ID][0]
        if not data:
            print(f"Failed to get constant for {self.Uni_ID}. Please add it to university_constants.csv first.")
            return

        # write constants to program_details.xlsx
        for header in headers:
            # Only write the constant if it's in headers
            if header in existing_headers:
                col_index = existing_headers.index(header) + 1
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row=row, column=col_index,
                               value=data[headers.index(header)])

        # Save the file
        try:
            wb.save(self.details_path)
            wb.save(self.details_path)
        except PermissionError:
            # save the file with a time stamp: YYYY-MM-DD_HH-MM-SS
            wb.save(os.path.join(
                f'data/{self.Uni_ID}/program_details_{time.strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'))
            print(
                f"PermissionError: failed to save program_details.xlsx. Saved as program_details_{time.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx instead.")

    def translate(self):
        # open program_details.xlsx
        wb = self.read_excel(self.details_path)
        sheet = wb.active

        # 获取现有的headers
        existing_headers = [cell.value for cell in sheet[1]]

        # 找出特定列名所在的列号
        col_institute_idx = existing_headers.index(header.college) + 1
        col_program_idx = existing_headers.index('专业') + 1
        col_background_idx = existing_headers.index(
            header.background_requirements) + 1

        # 翻译特定列并写入相应的中文列
        for row in range(2, sheet.max_row + 1):
            institute = sheet.cell(row=row, column=col_institute_idx).value
            program = sheet.cell(row=row, column=col_program_idx).value
            background = sheet.cell(row=row, column=col_background_idx).value

            sheet.cell(row=row, column=col_institute_idx +
                       1, value=translate_text(institute))
            sheet.cell(row=row, column=col_program_idx +
                       1, value=translate_text(program))
            sheet.cell(row=row, column=col_background_idx +
                       1, value=translate_text(background))

        # Save the file
        try:
            wb.save(self.details_path)
        except PermissionError:
            # save the file with a time stamp: YYYY-MM-DD_HH-MM-SS
            wb.save(os.path.join(
                f'data/{self.Uni_ID}/program_details_{time.strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'))
            print(
                f"PermissionError: failed to save program_details.xlsx. Saved as program_details_{time.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx instead.")

    def scrape_program_details(self):
        # 从 program_detail_headers.txt 读取标题
        # with open('data/program_detail_headers.txt', 'r', encoding='utf-8') as file:
        #     headers = [line.strip() for line in file.readlines()]
        # 获取header的所有自定义成员变量的值
        headers = header.ordered_list

        wb = self.read_excel(self.useful_link_path)
        sheet = wb.active

        new_wb = Workbook()
        new_sheet = new_wb.active
        new_sheet.append(headers)

        lock = threading.Lock()
        threads = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if self.test:
                # Execute the task without threading if self.test is True
                self.thread_task(row, headers, new_sheet, lock)
            else:   
                t = threading.Thread(target=self.thread_task,
                                    args=(row, headers, new_sheet, lock))
                threads.append(t)
                t.start()
        if not self.test:
            for t in threads:
                t.join()
        try:
            new_wb.save(self.details_path)
        except PermissionError:
            # save the file with a time stamp: YYYY-MM-DD_HH-MM-SS
            new_wb.save(os.path.join(
                f'data/{self.Uni_ID}/program_details_{time.strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'))
            print(
                f"PermissionError: failed to save program_details.xlsx. Saved as program_details_{time.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx instead.")

    def get_data_from_main_url(self, program_url, program_details):
        # try:
        #     response = requests.get(program_url)
        # except requests.ConnectionError as e:
        #     print(f"ConnectionError for URL {program_url}: {e}")
        #     return
        response = request_with_retry(
            url=program_url,
            Uni_ID=self.Uni_ID
        )
        if not response:
            print(f"ERROR: program_url: {program_url} returned badly!!!")
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        # self.get_program_intro(soup, program_details)
        self.get_background_requirements(soup, program_details, response.text)
        self.get_course_intro_and_details(soup, program_details, response.text)
        self.get_enrollment_deadlines(soup, program_details, response.text)
        self.get_tuition(soup, program_details, response.text)
        self.get_period(soup, program_details, response.text)
        self.get_language_requirements(soup, program_details, response.text)
        self.get_interview_requirements(soup, program_details, response.text)
        self.get_work_experience_requirements(
            soup, program_details, response.text)
        self.get_portfolio_requirements(soup, program_details, response.text)
        self.get_GRE_GMAT_requirements(soup, program_details, response.text)
        self.get_major_requirements_for_chinese_students(
            soup, program_details, response.text)
        self.get_major_requirements_for_uk_students(
            soup, program_details, response.text)
        self.get_major_specifications(soup, program_details, response.text)
        self.get_program_description(soup, program_details, response.text)
        self.get_application_deadlines(soup, program_details, response.text)

    def get_data_from_url(self, useful_links, program_details):
        # Here you can add more methods to process the soup based on the type of link_name.

        for link in useful_links:
            link_name, url = link.split(':')
            response = requests.get(url)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')
            # Assuming you will have methods like 'get_deadlines' to process the soup.
            # For example:
            # if "deadline" in link_name.lower():
            #     self.get_deadlines(soup, program_details)

    # Add your methods like get_deadlines, get_language_requirements, etc...

    # def get_program_intro(self, soup, program_details, extra_data=None):
    #     program_details[header.project_intro] = program_details[header.project_intro]

    def get_background_requirements(self, soup, program_details, extra_data=None):
        pass

    def get_course_intro_and_details(self, soup, program_details, extra_data=None):
        pass

    def get_enrollment_deadlines(self, soup, program_details, extra_data=None):
        pass

    def get_tuition(self, soup, program_details, extra_data=None):
        pass

    def get_period(self, soup, program_details, extra_data=None):
        pass

    def get_language_requirements(self, soup, program_details, extra_data=None):
        pass

    def _judge_interview_preference(self, text):
        required_phrases = ['may be', 'may also be',
                            'may be considered', 'may also be considered']
        for phrase in required_phrases:
            if phrase in text:
                return "可能要求"
        return "需要"

    def get_interview_requirements(self, soup, program_details, extra_data=None):
        # 定义关键词列表
        keywords = ['interview', 'special qualifying examination', 'qualifying essay',
                    'qualifying assessment', 'written examination', 'oral examination',
                    'oral test', 'required to pass a test', 'written examination', 'written test', 'written assessment']

        combined_text = ""
        # 搜索包含这些关键词的<p>标签
        program_description = soup.find(id="proxy_collapseprogramme")
        if program_description:
            for p_tag in program_description.find_all('p'):
                combined_text += p_tag.get_text().lower() + " "

        if program_details[header.background_requirements] != "未找到Entry requirements标签":
            combined_text += program_details[header.background_requirements].lower()

        combined_text = self._extract_relevant_text(combined_text, keywords)

        # 将合并后的文本添加到program_details字典中
        if combined_text:
            program_details[header.exam_requirements] = self._judge_interview_preference(
                combined_text)
            program_details[header.exam_requirements_details] = combined_text
        else:
            program_details[header.exam_requirements] = "未要求"

    def _judge_portfolio_preference(self, text):
        required_phrases = ['may be', 'may also be',
                            'may be considered', 'may also be considered']
        for phrase in required_phrases:
            if phrase in text:
                return "加分项"
        return "需要"

    def get_portfolio_requirements(self, soup, program_details, extra_data=None):
        phrases = ['written work', 'portfolio']
        # program_entry_req = soup.find(id="proxy_collapseentry_req")
        combined_text = ""

        if program_details[header.background_requirements] != "未找到Entry requirements标签":
            combined_text += program_details[header.background_requirements].lower()

            combined_text = self._extract_relevant_text(combined_text, phrases)
            if combined_text:
                program_details[header.portfolio] = self._judge_portfolio_preference(
                    combined_text)
                program_details[header.portfolio_details] = combined_text
            else:
                program_details[header.portfolio] = "未要求"
        else:
            program_details[header.portfolio] = "未找到Entry requirements标签"

    def _extract_relevant_text(self, text, keywords):
        # Extract sentences containing the phrase from the entry requirements section
        sentences = []
        line_sentences = re.split(r'(?<=[.!?])\s+', text)
        found = False
        for keyword in keywords:
            for sentence in line_sentences:
                if keyword in sentence:
                    sentences.append(sentence)
                    found = True
            # Avoid adding duplicate sentences when multiple keywords are matched
            if found:
                break

        # Combine the extracted sentences into one
        combined_text = ' '.join(sentences).strip()

        return combined_text

    def _judge_wrk_exp_preference(self, text):
        required_phrases = ['minimum of', 'at least', 'Ideally', 'is essential', 'must have', 'normally have',
                            'should also have', 'should have', 'need to have']
        for phrase in required_phrases:
            if phrase in text:
                return "需要"
        return "加分项"

    def get_work_experience_requirements(self, soup, program_details, extra_data=None):
        phrases = ['work experience', 'professional experience', 'industrial experience',
                   'existing engineering and design skills', 'practical experience',
                   'experience as a professional', ' who can demonstrate aptitude and experience',
                   'years of relevant experience', 'clinical experience', 'teaching experience',
                   'years\' experience working', 'years of training in',
                   'extensive experience', 'relevant experience',
                   'relevant quantitative or qualitative research experience',
                   'experience working', 'professional involvement',
                   'with equivalent experience', 'industry experience', 'relevant work',
                   'field experience', 'relevant employment']

        combined_text = ""

        if program_details[header.background_requirements] != "未找到Entry requirements标签":
            combined_text += program_details[header.background_requirements].lower()

            combined_text = self._extract_relevant_text(combined_text, phrases)
            if combined_text:
                program_details[header.work_experience_years] = self._judge_wrk_exp_preference(
                    combined_text)
                program_details[header.work_experience_details] = combined_text
            else:
                program_details[header.work_experience_years] = "未要求"
        else:
            program_details[header.work_experience_years] = "未找到Entry requirements标签"

    def get_GRE_GMAT_requirements(self, soup, program_details, extra_data=None):
        pass

    def get_major_requirements_for_chinese_students(self, soup, program_details, extra_data=None):
        pass

    def get_major_requirements_for_uk_students(self, soup, program_details, extra_data=None):
        pass

    def get_major_specifications(self, soup, program_details, extra_data=None):
        pass

    def get_program_description(self, soup, program_details, extra_data=None):
        pass

    def get_application_deadlines(self, soup, program_details, extra_data=None):
        pass

    # def get_campus(self, soup, program_details):
    #     pass

    # def British_local_requirements_for_display(self, soup, program_details):
    #     pass


