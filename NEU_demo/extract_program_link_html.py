import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# 读取 program_name_bank.xlsx 文件
wb = load_workbook('./data/program_word_bank.xlsx')
sheet = wb.active

# 创建program名字的列表
program_words = [row[0].lower() for row in sheet.iter_rows(min_row=2, values_only=True)]

# 获取HTML源代码
url = "https://graduateprograms.brown.edu/graduate_programs"
response = requests.get(url)
response.raise_for_status()  # 如果请求失败，这将引发异常

# 解析HTML
soup = BeautifulSoup(response.text, 'html.parser')

# 提取所有的 a 标签并检查链接名是否在 program_name_bank.xlsx 中
program_url_pairs = {}
for link in soup.find_all('a'):
    link_name = link.get_text().strip().lower()
    link_url = link.get('href')
    # 如果链接名在 program_name_bank.xlsx 中
    for word in program_words:
        if word in link_name:
            program_url_pairs[link_name] = link_url

# 打印结果
for program_name, url in program_url_pairs.items():
    print(f"Program name: {program_name}, URL: {url}")
