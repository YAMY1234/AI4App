import requests
import re
from openpyxl import load_workbook, Workbook
from bs4 import BeautifulSoup

def get_deadlines(soup, link_name, program_details):
    pass

def get_language_requirements(soup, link_name, program_details):
    pass

def get_program_courses(soup, link_name, program_details):
    pass

def get_GPA_wes_requirements(soup, link_name, program_details):
    pass

def get_GRE_requirements(soup, link_name, program_details):
    pass

def get_SAT_ACT_requirements(soup, link_name, program_details):
    pass

def get_recommendation_requirements(soup, link_name, program_details):
    pass

def get_essay_requirements(soup, link_name, program_details):
    pass

def get_other_requirements(soup, link_name, program_details):
    pass

def get_data_from_url(useful_links, program_details):

    # 遍历每个有用的链接
    for link in useful_links:
        # 分割 link 为 link_name 和 link_url
        link_name, url = link.split(':')

        # 发送 HTTP 请求
        response = requests.get(url)
        response.raise_for_status()  # 如果请求失败，这将引发异常

        # 使用 BeautifulSoup 解析 HTML 文档
        soup = BeautifulSoup(response.text, 'html.parser')

        # 使用正则表达式或其他方法从 soup 中提取需要的信息
        get_deadlines(soup, link_name, program_details)


    return "Some Data"


def main():
    # 读取 program_useful_link.xlsx 文件
    wb = load_workbook('./data/program_useful_link.xlsx')
    sheet = wb.active

    # 创建新的 Excel 文件并添加标题行
    new_wb = Workbook()
    new_sheet = new_wb.active
    new_sheet.append(["Program", "DDL - early action", "DDL - early decision", "DDL - regular action",
                      "DDL - regular decision", "GPA要求", "wes要求", "推荐信要求（份数）", "托福", "雅思",
                      "多邻国", "GRE", "SAT", "ACT", "是否面试", "文书要求（text）"])

    # 遍历每一行，获取 program name 和 useful_links
    for row in sheet.iter_rows(min_row=2, values_only=True):
        program_name, intro_text, *useful_links = row

        # 创建一个字典存储每个项目的详细信息
        program_details = {}

        # 获取链接中的数据，直接修改program_details当中的内容
        get_data_from_url(useful_links, program_details)

        # 添加数据到新的 Excel 文件中
        new_sheet.append([program_name, program_details.get("DDL - early action"), program_details.get("DDL - early decision"),
                          program_details.get("DDL - regular action"), program_details.get("DDL - regular decision"),
                          program_details.get("GPA要求"), program_details.get("wes要求"), program_details.get("推荐信要求（份数）"),
                          program_details.get("托福"), program_details.get("雅思"), program_details.get("多邻国"),
                          program_details.get("GRE"), program_details.get("SAT"), program_details.get("ACT"),
                          program_details.get("是否面试"), program_details.get("文书要求（text）")])

    # 保存新的 Excel 文件
    new_wb.save("./data/program_details.xlsx")


if __name__ == "__main__":
    main()
