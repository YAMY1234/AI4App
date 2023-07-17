import json
from openpyxl import Workbook, load_workbook

# 读取program-wordBank.xlsx文件中的程序单词
word_bank_file = './data/program-wordBank.xlsx'
word_bank_workbook = load_workbook(word_bank_file)
word_bank_sheet = word_bank_workbook.active

program_words = []
for row in word_bank_sheet.iter_rows(min_row=2, max_col=1, values_only=True):
    if row[0]:
        program_words.append(row[0])

# 读取NEU_demo.json文件
json_file = './data/NEU_demo.json'
with open(json_file, 'r') as file:
    json_data = json.load(file)

# 创建用于存储program name和URL的字典
program_url_dict = {}

# 创建一个函数来递归遍历json_data
def search_json(obj):
    program_name = None
    url = None

    if isinstance(obj, dict):
        for k, v in obj.items():
            if isinstance(v, (list, dict)):
                child_program_name, child_url = search_json(v)
                if child_program_name is not None:
                    program_name = child_program_name
                if child_url is not None:
                    url = child_url
            elif isinstance(v, str):
                if any(word in v for word in program_words):
                    program_name = v
                elif 'http' in v:
                    url = v
        if program_name is not None and url is None:
            for k, v in obj.items():
                if isinstance(v, str) and 'http' in v:
                    url = v
        if program_name is not None and url is not None:
            program_url_dict[program_name] = url
    elif isinstance(obj, list):
        for item in obj:
            child_program_name, child_url = search_json(item)
            if child_program_name is not None:
                program_name = child_program_name
            if child_url is not None:
                url = child_url

    return program_name, url

search_json(json_data)

# 创建Excel工作簿和工作表
output_file = './data/program_url_pair.xlsx'
workbook = Workbook()
sheet = workbook.active

# 写入标题行
sheet.append(['Program Name', 'URL'])

# 写入program name和URL
for program_name, url in program_url_dict.items():
    sheet.append([program_name, url])

# 保存Excel文件
workbook.save(output_file)
