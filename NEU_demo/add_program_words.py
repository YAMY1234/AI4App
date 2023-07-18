import os
from openpyxl import Workbook, load_workbook

# 读取txt文件并提取长度大于等于4的词汇
def read_txt_file(file_path):
    with open(file_path, 'r') as f:
        words = f.read().split()
    words = [word for word in words if len(word) >= 4]
    return words

# 读取xlsx文件的现有词汇
def read_xlsx_file(file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    words = [cell.value for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1) for cell in row]
    return words

# 对词汇进行排序并写入xlsx文件
def write_xlsx_file(file_path, words):
    workbook = Workbook()
    sheet = workbook.active
    for word in sorted(words):
        sheet.append([word])
    workbook.save(file_path)

# 读取txt文件中的词汇
txt_file_path = './data/program_additional.txt'
txt_words = read_txt_file(txt_file_path)

# 读取xlsx文件中的词汇
xlsx_file_path = './data/program_word_bank.xlsx'
xlsx_words = read_xlsx_file(xlsx_file_path)

# 判断txt文件中的词汇是否在xlsx文件中，如果不在则添加
txt_words = set(txt_words)
txt_words.remove("with")
txt_words.remove("over")
for word in txt_words:
    if word not in xlsx_words:
        xlsx_words.append(word)

# 将更新后的词汇列表写入xlsx文件
write_xlsx_file(xlsx_file_path, xlsx_words)
