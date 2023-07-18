import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook
import re
import concurrent.futures

# 读取 program_url_pair.xlsx 文件
wb = load_workbook('./data/program_url_pair.xlsx')
link_bank = load_workbook('./data/url_bank.xlsx').active
sheet = wb.active

# 创建新的 Excel 文件
new_wb = Workbook()
new_sheet = new_wb.active

# 添加标题行
new_sheet.append(['项目名', '项目简介', '链接'])

def process_row(row):
    print("processing:" + row[0])
    program_name, url = row

    # 发送 HTTP 请求
    response = requests.get(url)
    response.raise_for_status()  # 如果请求失败，这将引发异常

    # 使用 BeautifulSoup 解析 HTML 文档
    soup = BeautifulSoup(response.text, 'html.parser')

    # 提取专业介绍
    intro_div = soup.find('div', class_='zero')  # 根据实际情况可能需要调整   [, class_='zero']
    if intro_div is not None:
        intro_text = intro_div.get_text()

    # 提取有用的子链接 URL
    useful_links = []
    for link in soup.find_all('a'):
        link_name = link.get_text().strip()
        link_url = link.get('href')
        # 如果 link_name 在 url_bank.xlsx 文件的第一列中
        for bank_row in link_bank.iter_rows(min_row=2, values_only=True):
            if bank_row[0].lower() in link_name.lower():
                useful_links.append(f'{link_name}:{link_url}')
                break

    return [program_name, intro_text] + useful_links

# 使用 ThreadPoolExecutor 创建线程池并行处理每一行
with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
    results = list(executor.map(process_row, sheet.iter_rows(min_row=2, values_only=True)))

# 将结果写入新的 Excel 文件
for result in results:
    new_sheet.append(result)

# 保存新的 Excel 文件
new_wb.save('./data/program_useful_link.xlsx')
